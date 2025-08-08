# from seleniumbase import SB

# with SB(uc=True, test=True) as sb:
#     url = "https://2captcha.com/demo/cloudflare-turnstile"
#     sb.uc_open_with_reconnect(url, reconnect_time=2)
#     sb.uc_gui_handle_captcha()
#     # sb.assert_element("img#captcha-success", timeout=3)
#     # sb.set_messenger_theme(location="top_left")
#     # sb.post_message("SeleniumBase wasn't detected", duration=3)
import openpyxl

def get_column_values_without_header(file_path, sheet_name, column_index):
    """
    Reads an XLSX file, extracts values from a specified column 
    (excluding the header row if present), and returns them as a list.

    Args:
        file_path (str): The path to the XLSX file.
        sheet_name (str): The name of the worksheet to read.
        column_index (int): The 1-based index of the column to extract 
                            (e.g., 1 for column A, 2 for column B).

    Returns:
        list: A list containing the values from the specified column,
              excluding the header row.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)  # Load the Excel workbook
        worksheet = workbook[sheet_name]  # Select the worksheet by name
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
        return []
    except KeyError:
        print(f"Error: The sheet '{sheet_name}' was not found in the workbook.")
        return []

    column_values = []
    # Iterate through rows starting from the second row (skipping the header)
    for row_num in range(2, worksheet.max_row + 1):  # Starting from row 2 (0-indexed Python)
        cell_value = worksheet.cell(row=row_num, column=column_index).value  # Access cell value
        column_values.append(cell_value)

    return column_values

def write_to_cell(file_path, sheet_name, row, column, value):
    """
    Writes a value to a specific cell in an XLSX file.

    Args:
        file_path (str): The path to the XLSX file.
        sheet_name (str): The name of the worksheet to modify.
        row (int): The row number of the cell (1-based index).
        column (int): The column number of the cell (1-based index).
        value: The value to write to the cell.
    """
    try:
        workbook = openpyxl.load_workbook(file_path) # Load the existing workbook
        worksheet = workbook[sheet_name] # Select the worksheet by name

        # Write the value to the specified cell using row and column numbers
        worksheet.cell(row=row, column=column, value=value)

        workbook.save(file_path) # Save the modified workbook
        print(f"Value '{value}' written to cell {chr(64 + column)}{row} in sheet '{sheet_name}'")
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
    except KeyError:
        print(f"Error: The sheet '{sheet_name}' was not found in the workbook.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
file = "./kaufland.xlsx" # Replace with your file name
sheet = "Sheet1" # Replace with your sheet name
column = 2 # Column B (1-based index)

data_list = get_column_values_without_header(file, sheet, column)

if data_list:
    data_list.insert(0, 'dummy_1')
    data_list.insert(1, 'dummy_2')
    start = input("Enter the starting index (default is 2): ")
    end = input("Enter the ending index (default all): ")
    start = int(start) if start else 2  # Default start index is 2 (to skip the header)
    end = int(end)+1 if end else -1
    eans = data_list[start:end]  # Slice the list based on user input
else:
    eans = []
    exit(0)

from seleniumbase import SB
import time
import re

def verify_success(sb):
    sb.assert_element('img[alt="Logo Assembly"]', timeout=4)
    sb.sleep(3)

with SB(uc=True,test=True) as sb:
    sb.activate_cdp_mode("https://www.kaufland.de")
    # sb.sleep(2)
    # if sb.is_text_visible("Verifizierung erforderlich","h1"):
    #     sb.sleep(2)
    #     sb.uc_gui_click_captcha()
    #     sb.sleep(3)
    #     sb.wait_for_text_not_visible("Verifizierung erforderlich","h1", timeout=30)
    # # try:
    # #     print("just verify")
    # #     verify_success(sb)
    # # except Exception:
    # print("mouse click")
    # while sb.is_text_visible("Verifizierung erforderlich","h1"):
    #     sb.wait(10)
    #     sb.uc_gui_click_captcha()
    #     sb.wait(10)
    #     break
    i=start
    for ean in eans:
        print(f"Row: {i}, Processing EAN: {ean}")
        sb.uc_open_with_reconnect(f"https://www.kaufland.de/api/search/v1/result-product-offers/?requestType=initial-load&productOffset=0&pageType=search&searchValue={ean}&deviceType=desktop&useNewUrls=true", 3)
        if sb.is_text_visible("Verifizierung erforderlich","h1"):
            sb.sleep(5)
            sb.uc_gui_click_captcha()
            sb.sleep(5)
            sb.wait_for_text_not_visible("Verifizierung erforderlich","h1", timeout=30)
        # time.sleep(1)  # Wait for the page to load
        page_source = sb.get_page_source()  # Get the page source
        # print(page_source)  # Print the page source to verify the content
        pattern = r'"id":(\d+)'  # Captures the digits after '"id":'
        match = re.search(pattern, page_source)

        if match:
            product_id = match.group(1)  # Extracts the captured group (the digits)
            len(product_id)
            if len(product_id) > 7:
                print(f"Product ID: {product_id}")
                print(f"https://www.kaufland.de/product/{product_id}/?search_value={ean}")
                if_url = f"https://www.kaufland.de/product/{product_id}/?search_value={ean}"
                file = "./kaufland.xlsx" # Replace with your Excel file name
                sheet = "Sheet1" # Replace with your sheet name
                row_num = i
                col_num = 3 # Column C

                write_to_cell(file, sheet, row_num, col_num, if_url)
        else:
            print("Product ID not found.")
        i += 1

        # exit(0)
        # cookie_element = sb.find_element("//div[@data-testid='product-tiles']", timeout=10)
        # if cookie_element:
        #     sb.click("//button[@id='onetrust-accept-btn-handler']")  # Click the link element
        # time.sleep(2)  # Wait for the page to load after clicking
        # link_element = sb.find_element("//div[@data-testid='product-tiles']", timeout=10)
        # if link_element:
        # # Extract the href attribute
        #     href_value = link_element.get_attribute("href")
        #     # Print the extracted href
        #     print(f"The extracted href is: {href_value}")

    # sb.assert_text("Verify you are human",timeout=10)
    # if sb.is_element_visible('input[value*="Verify"]'):
    #     sb.uc_click('input[value*="Verify"]')
    # else:
    #     print("function click")
    #     sb.uc_gui_click_captcha()
    #     sb.wait(120)